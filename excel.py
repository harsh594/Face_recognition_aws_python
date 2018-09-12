def excel(l):
 import openpyxl
 import os
 from datetime import date
 
 now=date.today()
 now=now.strftime('%d-%m-%Y')
 folderName = "Attendence"
 folderPath = os.path.join(os.path.dirname(os.path.realpath('__file__')), folderName)
 wb=openpyxl.Workbook()
 if not os.path.exists(folderPath):
    os.makedirs(folderPath)
 try:
  wb = openpyxl.load_workbook(folderPath+"\\"+"Attendence_Register.xlsx")
 except:
   wb.save(folderPath+"\\"+"Attendence_Register.xlsx") 
   wb = openpyxl.load_workbook(folderPath+"\\"+"Attendence_Register.xlsx")
 sheet = wb.active
 i=0
 while(True):
    x = sheet.cell(row = 1, column = i+1)
   
    j = 2
  
    if x.value == now:
        for a in range(len(l)):
            while(True):
             y=sheet.cell(row = j, column = i+1)   
             if y.value !=None and y.value!=l[a]:
              j+=1
             elif y.value==l[a]:
               j=2  
               break  
             else:
                 
               y.value = l[a]
               j=2
               break
        print("Old data updated") 
        wb.save(folderPath+"\\"+"Attendence_Register.xlsx")
        return    
    if x.value!=now and x.value is not None:
      i+=1  
    else:
        x.value = now
        for a in range(len(l)):
            y = sheet.cell(row = j, column = i+1)
            y.value = l[a]
            j+=1
        print("New data entered")
        wb.save(folderPath+"\\"+"Attendence_Register.xlsx")
        return 
if __name__=='__main__':
 excel(l)
 